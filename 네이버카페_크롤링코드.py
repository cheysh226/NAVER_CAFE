# %%
#########################################
keyword = ['영유아', '역량강화']
#########################################
## 0. 초기 셋팅 부분 (라이브러리, driver 시작)

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl.styles import *
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
import datetime
import os

def sanitize_excel_value(value: str) -> str:
	"""엑셀에서 수식으로 인식될 가능성이 있는 값을 정리"""
	if isinstance(value, str) and re.match(r"^[=\+\-@]", value):
		return value[1:]  # 첫 글자 제거
	return value

driver = webdriver.Chrome()
driver.maximize_window()


# %%
## 1. 엑셀 양식 만들기
wb = openpyxl.Workbook()
ws = wb.active

# 제목 적기
sub = ['분류', '작성일자', '카페글 제목', '내용', '카페 주소', '카페명','이하 댓글']
for kwd, j in zip(sub, list(range(1, len(sub)+1))):
	ws.cell(row=1, column=j).value = kwd
for cell in ws['1']:
	cell.font = Font(bold=True)
	cell.border = Border(bottom=Side(style='double'))
ws.freeze_panes = 'A2' # 첫행고정

# 셀너비
ws.column_dimensions['A'].width = 6 # A열
ws.column_dimensions['B'].width = 12 # B열
ws.column_dimensions['C'].width = 55 # C열
ws.column_dimensions['D'].width = 60

# 셀 1행 배경색 노랑색 변경
y_color = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
for num in range(1, len(sub)+1):
	ws.cell(1,num).fill = y_color
# 가운데 정렬
for num in range(1, len(sub)+1):
	for row in range(1, len(ws['A'])+1):
		ws.cell(row=row, column=num).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
i=2


# %%
## 3. 크롤링 부분 --- 10분 정도 소요됨
driver.get(f'https://search.naver.com/search.naver?ssc=tab.cafe.all&sm=tab_jum&query={"+".join(keyword)}')

titleElements = driver.find_elements(By.CSS_SELECTOR,'div.title_area > a')
for titleElem in titleElements:
	time.sleep(1)
	cafelink = titleElem.get_attribute('href')
	if 'cafe.naver.com' not in cafelink:
		continue
	driver.switch_to.new_window('tab')
	driver.get(cafelink)
	time.sleep(1)
	cafeName = driver.find_element(By.CSS_SELECTOR,'h1').text.strip()
	driver.switch_to.default_content()
	driver.switch_to.frame(driver.find_element(By.CSS_SELECTOR,'#cafe_main'))
	try:
		title = driver.find_element(By.CSS_SELECTOR,'h3.title_text').text.strip()
		context = driver.find_element(By.CSS_SELECTOR,'div.content.CafeViewer').text.strip()
		created_at = driver.find_element(By.CSS_SELECTOR,'div.article_header span.date').text.strip()
		print(title, context, created_at)
	except:
		pass
	created_at = datetime.datetime.strptime(created_at,'%Y.%m.%d. %H:%M').strftime('%Y-%m-%d')

	title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
	context = ILLEGAL_CHARACTERS_RE.sub(r'', context)
	title = sanitize_excel_value(title)
	context = sanitize_excel_value(context)
	# ['분류', '작성일자', '카페글 제목', '내용', '카페 주소','','카페글 이모지제거','내용 이모지제거', '카페명','이하 댓글']
	# 댓글들
	replies = [commentElem.find_element(By.CSS_SELECTOR,"span.text_comment").text for commentElem in driver.find_elements(By.CSS_SELECTOR,'li.CommentItem')]

	# tr = ['cafe', created_at, title, context, cafelink,'',remove_emoji(title), remove_emoji(context),cafeName ] + replies
	tr = ['cafe', created_at, title, context, cafelink,cafeName ] + replies

	ws.append(tr)

	driver.close()
	driver.switch_to.window(driver.window_handles[0])

# 파일 저장
filename = f"{'_'.join(keyword)}_네이버카페.xlsx"
wb.save(filename)

# 엑셀 파일 자동으로 열기 (Windows용)
os.startfile(filename)


